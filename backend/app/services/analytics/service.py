"""AnalyticsService — period-aware aggregated analytics for the back-office.

Computes KPIs with trends, time series, language distribution,
question type breakdown, top questions, and exports.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO

import structlog
from sqlalchemy import Date, case, cast, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.tenant import TenantContext
from app.models.contact import Contact
from app.models.conversation import Conversation, Message
from app.models.enums import (
    ConversationStatus,
    FeedbackRating,
    UnansweredStatus,
)
from app.models.escalation import Escalation
from app.models.feedback import Feedback, UnansweredQuestion

logger = structlog.get_logger()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _resolve_period(
    period: str,
    start: str | None,
    end: str | None,
) -> tuple[datetime, datetime, datetime, datetime]:
    """Convert period string to (cur_start, cur_end, prev_start, prev_end)."""
    now = datetime.now(UTC)

    if period == "custom" and start and end:
        cur_start = datetime.fromisoformat(start).replace(tzinfo=UTC)
        cur_end = datetime.fromisoformat(end).replace(tzinfo=UTC)
        delta = cur_end - cur_start
        prev_start = cur_start - delta
        prev_end = cur_start
    else:
        days_map = {"7d": 7, "30d": 30, "90d": 90}
        days = days_map.get(period, 30)
        cur_end = now
        cur_start = now - timedelta(days=days)
        prev_end = cur_start
        prev_start = prev_end - timedelta(days=days)

    return cur_start, cur_end, prev_start, prev_end


def _trend(current: float, previous: float) -> float:
    """Compute percentage change, rounded to 1 decimal."""
    if previous == 0:
        return 0.0
    return round(((current - previous) / previous) * 100, 1)


def _fill_date_gaps(
    data: dict[str, dict],
    start: datetime,
    end: datetime,
    defaults: dict,
) -> list[dict]:
    """Fill missing dates with zero-value entries."""
    result: list[dict] = []
    current = start.date()
    end_date = end.date()
    while current <= end_date:
        key = current.isoformat()
        if key in data:
            result.append({"date": key, **data[key]})
        else:
            result.append({"date": key, **defaults})
        current += timedelta(days=1)
    return result


LANGUAGE_LABELS: dict[str, str] = {
    "fr": "Français",
    "ar": "العربية",
    "en": "English",
}


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------


class AnalyticsService:
    """Compute analytics for a tenant within a given period."""

    def __init__(self) -> None:
        self._logger = logger.bind(service="analytics_service")

    # ── Overview KPIs ────────────────────────────────────────────────────

    async def get_overview(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
    ) -> dict:
        cur_start, cur_end, prev_start, prev_end = _resolve_period(period, start, end)

        async with tenant.db_session() as session:
            # Conversations
            cur_conv = await self._count(
                session,
                Conversation.id,
                Conversation.started_at,
                cur_start,
                cur_end,
            )
            prev_conv = await self._count(
                session,
                Conversation.id,
                Conversation.started_at,
                prev_start,
                prev_end,
            )

            # Messages
            cur_msg = await self._count(
                session,
                Message.id,
                Message.timestamp,
                cur_start,
                cur_end,
            )
            prev_msg = await self._count(
                session,
                Message.id,
                Message.timestamp,
                prev_start,
                prev_end,
            )

            # Resolution rate
            cur_resolution = await self._resolution_rate(session, cur_start, cur_end)
            prev_resolution = await self._resolution_rate(session, prev_start, prev_end)

            # CSAT
            cur_csat = await self._csat_score(session, cur_start, cur_end)
            prev_csat = await self._csat_score(session, prev_start, prev_end)

        return {
            "conversations_total": cur_conv,
            "conversations_trend": _trend(cur_conv, prev_conv),
            "messages_total": cur_msg,
            "messages_trend": _trend(cur_msg, prev_msg),
            "resolution_rate": cur_resolution,
            "resolution_trend": _trend(cur_resolution, prev_resolution),
            "csat_average": cur_csat,
            "csat_trend": _trend(cur_csat, prev_csat),
        }

    # ── Time series ──────────────────────────────────────────────────────

    async def get_timeseries(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
    ) -> list[dict]:
        cur_start, cur_end, _, _ = _resolve_period(period, start, end)

        async with tenant.db_session() as session:
            # Daily conversations
            conv_q = await session.execute(
                select(
                    cast(Conversation.started_at, Date).label("day"),
                    func.count(Conversation.id).label("cnt"),
                )
                .where(Conversation.started_at.between(cur_start, cur_end))
                .group_by("day")
            )
            conv_by_day = {str(row.day): row.cnt for row in conv_q.all()}

            # Daily messages
            msg_q = await session.execute(
                select(
                    cast(Message.timestamp, Date).label("day"),
                    func.count(Message.id).label("cnt"),
                )
                .where(Message.timestamp.between(cur_start, cur_end))
                .group_by("day")
            )
            msg_by_day = {str(row.day): row.cnt for row in msg_q.all()}

            # Daily escalations
            esc_q = await session.execute(
                select(
                    cast(Escalation.created_at, Date).label("day"),
                    func.count(Escalation.id).label("cnt"),
                )
                .where(Escalation.created_at.between(cur_start, cur_end))
                .group_by("day")
            )
            esc_by_day = {str(row.day): row.cnt for row in esc_q.all()}

        # Merge into gap-filled daily series
        merged: dict[str, dict] = {}
        all_days = set(conv_by_day) | set(msg_by_day) | set(esc_by_day)
        for day in all_days:
            merged[day] = {
                "conversations": conv_by_day.get(day, 0),
                "messages": msg_by_day.get(day, 0),
                "escalations": esc_by_day.get(day, 0),
            }

        return _fill_date_gaps(
            merged,
            cur_start,
            cur_end,
            {"conversations": 0, "messages": 0, "escalations": 0},
        )

    # ── Language distribution ────────────────────────────────────────────

    async def get_languages(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
    ) -> list[dict]:
        cur_start, cur_end, _, _ = _resolve_period(period, start, end)

        async with tenant.db_session() as session:
            result = await session.execute(
                select(
                    Contact.language,
                    func.count(Conversation.id).label("cnt"),
                )
                .join(Conversation, Conversation.contact_id == Contact.id)
                .where(Conversation.started_at.between(cur_start, cur_end))
                .group_by(Contact.language)
            )
            rows = result.all()

        total = sum(r.cnt for r in rows) or 1
        return [
            {
                "language": r.language or "fr",
                "label": LANGUAGE_LABELS.get(r.language or "fr", r.language or "fr"),
                "count": r.cnt,
                "percentage": round(r.cnt / total * 100, 1),
            }
            for r in rows
        ]

    # ── Question type distribution ───────────────────────────────────────

    async def get_question_types(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
    ) -> list[dict]:
        cur_start, cur_end, _, _ = _resolve_period(period, start, end)

        async with tenant.db_session() as session:
            result = await session.execute(
                select(
                    Conversation.status,
                    func.count(Conversation.id).label("cnt"),
                )
                .where(Conversation.started_at.between(cur_start, cur_end))
                .group_by(Conversation.status)
            )
            rows = {r.status: r.cnt for r in result.all()}

        faq = rows.get(ConversationStatus.ended, 0)
        escalated = rows.get(ConversationStatus.escalated, 0) + rows.get(
            ConversationStatus.human_handled, 0
        )
        active = rows.get(ConversationStatus.active, 0)
        total = faq + escalated + active or 1

        items = [
            {"type": "faq", "label": "FAQ / Résolu", "count": faq},
            {"type": "escalade", "label": "Escaladé", "count": escalated},
            {"type": "en_cours", "label": "En cours", "count": active},
        ]
        for item in items:
            item["percentage"] = round(item["count"] / total * 100, 1)
        return items

    # ── Top questions ────────────────────────────────────────────────────

    async def get_top_questions(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
        limit: int = 10,
    ) -> list[dict]:
        cur_start, cur_end, _, _ = _resolve_period(period, start, end)

        async with tenant.db_session() as session:
            result = await session.execute(
                select(
                    UnansweredQuestion.question,
                    UnansweredQuestion.frequency,
                    UnansweredQuestion.status,
                )
                .where(UnansweredQuestion.created_at.between(cur_start, cur_end))
                .order_by(UnansweredQuestion.frequency.desc())
                .limit(limit)
            )
            rows = result.all()

        covered_statuses = {UnansweredStatus.approved, UnansweredStatus.injected}
        return [
            {
                "question": r.question,
                "count": r.frequency,
                "avg_confidence": 0.0,
                "status": "covered" if r.status in covered_statuses else "uncovered",
            }
            for r in rows
        ]

    # ── Exports ──────────────────────────────────────────────────────────

    async def export_excel(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
    ) -> BytesIO:
        from openpyxl import Workbook

        overview = await self.get_overview(tenant, period, start, end)
        timeseries = await self.get_timeseries(tenant, period, start, end)
        languages = await self.get_languages(tenant, period, start, end)
        top_q = await self.get_top_questions(tenant, period, start, end, limit=50)

        wb = Workbook()

        # Sheet 1: Overview
        ws = wb.active
        ws.title = "Vue d'ensemble"
        ws.append(["Métrique", "Valeur", "Tendance (%)"])
        ws.append(
            ["Conversations", overview["conversations_total"], overview["conversations_trend"]]
        )
        ws.append(["Messages", overview["messages_total"], overview["messages_trend"]])
        ws.append(
            ["Taux résolution (%)", overview["resolution_rate"], overview["resolution_trend"]]
        )
        ws.append(["CSAT moyen (/5)", overview["csat_average"], overview["csat_trend"]])

        # Sheet 2: Time Series
        ws2 = wb.create_sheet("Séries temporelles")
        ws2.append(["Date", "Conversations", "Messages", "Escalades"])
        for pt in timeseries:
            ws2.append([pt["date"], pt["conversations"], pt["messages"], pt["escalations"]])

        # Sheet 3: Languages
        ws3 = wb.create_sheet("Langues")
        ws3.append(["Langue", "Conversations", "Pourcentage (%)"])
        for lang in languages:
            ws3.append([lang["label"], lang["count"], lang["percentage"]])

        # Sheet 4: Top Questions
        ws4 = wb.create_sheet("Questions fréquentes")
        ws4.append(["Question", "Fréquence", "Statut"])
        for q in top_q:
            ws4.append([q["question"], q["count"], q["status"]])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    async def export_pdf(
        self,
        tenant: TenantContext,
        period: str,
        start: str | None = None,
        end: str | None = None,
    ) -> BytesIO:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        overview = await self.get_overview(tenant, period, start, end)
        top_q = await self.get_top_questions(tenant, period, start, end, limit=20)

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        elements: list = []

        elements.append(Paragraph(f"Analytics — CRI ({period})", styles["Title"]))
        elements.append(Spacer(1, 0.5 * cm))

        # KPIs table
        kpi_data = [
            ["Métrique", "Valeur", "Tendance"],
            [
                "Conversations",
                str(overview["conversations_total"]),
                f"{overview['conversations_trend']}%",
            ],
            ["Messages", str(overview["messages_total"]), f"{overview['messages_trend']}%"],
            [
                "Taux résolution",
                f"{overview['resolution_rate']}%",
                f"{overview['resolution_trend']}%",
            ],
            ["CSAT moyen", f"{overview['csat_average']}/5", f"{overview['csat_trend']}%"],
        ]
        kpi_table = Table(kpi_data, colWidths=[7 * cm, 5 * cm, 5 * cm])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C4704B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#FAF7F2")],
                    ),
                ]
            )
        )
        elements.append(kpi_table)
        elements.append(Spacer(1, 1 * cm))

        # Top questions table
        if top_q:
            elements.append(Paragraph("Questions fréquentes", styles["Heading2"]))
            elements.append(Spacer(1, 0.3 * cm))
            q_data = [["#", "Question", "Fréquence", "Statut"]]
            for i, q in enumerate(top_q, 1):
                question_text = (
                    q["question"][:80] + "…" if len(q["question"]) > 80 else q["question"]
                )
                q_data.append([str(i), question_text, str(q["count"]), q["status"]])
            q_table = Table(q_data, colWidths=[1 * cm, 10 * cm, 3 * cm, 3 * cm])
            q_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C4704B")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#FAF7F2")],
                        ),
                    ]
                )
            )
            elements.append(q_table)

        doc.build(elements)
        buf.seek(0)
        return buf

    # ── Private helpers ──────────────────────────────────────────────────

    @staticmethod
    async def _count(
        session: AsyncSession,
        column,
        date_column,
        start: datetime,
        end: datetime,
    ) -> int:
        result = await session.execute(
            select(func.count(column)).where(date_column.between(start, end))
        )
        return result.scalar_one()

    @staticmethod
    async def _resolution_rate(
        session: AsyncSession,
        start: datetime,
        end: datetime,
    ) -> float:
        result = await session.execute(
            select(
                func.count(Conversation.id).label("total"),
                func.count(
                    case(
                        (Conversation.status == ConversationStatus.ended, Conversation.id),
                    )
                ).label("ended"),
            ).where(Conversation.started_at.between(start, end))
        )
        row = result.one()
        if row.total == 0:
            return 0.0
        return round(row.ended / row.total * 100, 1)

    @staticmethod
    async def _csat_score(
        session: AsyncSession,
        start: datetime,
        end: datetime,
    ) -> float:
        result = await session.execute(
            select(
                Feedback.rating,
                func.count().label("cnt"),
            )
            .where(Feedback.created_at.between(start, end))
            .group_by(Feedback.rating)
        )
        counts = {row.rating: row.cnt for row in result.all()}
        positive = counts.get(FeedbackRating.positive, 0)
        negative = counts.get(FeedbackRating.negative, 0)
        denominator = positive + negative
        if denominator == 0:
            return 0.0
        return round((positive / denominator) * 5, 1)


# ---------------------------------------------------------------------------
# Singleton
# ---------------------------------------------------------------------------
_analytics_service: AnalyticsService | None = None


def get_analytics_service() -> AnalyticsService:
    """Get or create the AnalyticsService singleton."""
    global _analytics_service  # noqa: PLW0603
    if _analytics_service is None:
        _analytics_service = AnalyticsService()
    return _analytics_service
