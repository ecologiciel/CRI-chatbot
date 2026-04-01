"""DossierImportService — secure Excel/CSV import for dossier tracking.

Implements Level 1 SI integration (CPS R8): Excel/CSV export from CRI SI,
imported into PostgreSQL with validation, sanitisation, dedup, change
detection, and history tracking.

Security invariants:
- openpyxl ``data_only=True`` — never execute macros/formulas
- File extension whitelist + size cap
- HTML stripping + SQL injection pattern neutralisation on every field
- SHA-256 dedup against prior completed imports
- All DB access scoped to ``tenant.db_session()``
"""

from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import time
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import PurePosixPath
from typing import Any

import structlog
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.tenant import TenantContext
from app.models.contact import Contact
from app.models.dossier import Dossier, DossierHistory
from app.models.enums import DossierStatut, SyncStatus
from app.models.sync import SyncLog

logger = structlog.get_logger()

# ── Constants ────────────────────────────────────────────────────────

MAX_FILE_SIZE_MB: int = 10
MAX_ROWS: int = 50_000
BATCH_SIZE: int = 500
ALLOWED_EXTENSIONS: set[str] = {".xlsx", ".xls", ".csv"}

# Pre-compiled SQL injection patterns (defense-in-depth)
SQL_INJECTION_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"(?i)\bOR\s+1\s*=\s*1\b"),
    re.compile(r"(?i)\bUNION\s+SELECT\b"),
    re.compile(r"(?i);\s*DROP\b"),
    re.compile(r"(?i);\s*DELETE\b"),
    re.compile(r"(?i);\s*UPDATE\b"),
    re.compile(r"(?i)\bxp_cmdshell\b"),
    re.compile(r"--\s"),
]

# Statut string → DossierStatut enum (case-insensitive, accent-tolerant)
STATUT_MAPPING: dict[str, DossierStatut] = {
    "en cours": DossierStatut.en_cours,
    "en_cours": DossierStatut.en_cours,
    "valide": DossierStatut.valide,
    "validé": DossierStatut.valide,
    "rejete": DossierStatut.rejete,
    "rejeté": DossierStatut.rejete,
    "en attente": DossierStatut.en_attente,
    "en_attente": DossierStatut.en_attente,
    "complement": DossierStatut.complement,
    "complément": DossierStatut.complement,
    "demande de complement": DossierStatut.complement,
    "demande de complément": DossierStatut.complement,
    "incomplet": DossierStatut.incomplet,
}

# Column alias sets for auto-detection from French Excel headers
NUMERO_ALIASES: set[str] = {
    "numero", "numéro", "n°_dossier", "n°dossier", "num_dossier",
    "numero_dossier", "ref", "reference", "référence",
}
STATUT_ALIASES: set[str] = {"statut", "status", "état", "etat"}
TYPE_PROJET_ALIASES: set[str] = {
    "type_projet", "type_de_projet", "type", "projet",
}
RAISON_SOCIALE_ALIASES: set[str] = {
    "raison_sociale", "raison_social", "entreprise", "société", "societe",
    "nom_entreprise",
}
MONTANT_ALIASES: set[str] = {
    "montant_investissement", "montant", "investissement", "montant_mad",
}
REGION_ALIASES: set[str] = {"region", "région"}
SECTEUR_ALIASES: set[str] = {"secteur", "secteur_activité", "secteur_activite"}
DATE_DEPOT_ALIASES: set[str] = {
    "date_depot", "date_dépôt", "date_de_depot", "date_de_dépôt",
    "date_creation", "date_création",
}
OBSERVATIONS_ALIASES: set[str] = {
    "observations", "observation", "commentaire", "commentaires", "notes",
}
PHONE_ALIASES: set[str] = {
    "phone", "telephone", "téléphone", "tel", "mobile", "numéro_tel",
}

# Ordered list mapping alias sets to DossierImportRow field names
_FIELD_ALIAS_MAP: list[tuple[str, set[str]]] = [
    ("numero", NUMERO_ALIASES),
    ("statut", STATUT_ALIASES),
    ("type_projet", TYPE_PROJET_ALIASES),
    ("raison_sociale", RAISON_SOCIALE_ALIASES),
    ("montant_investissement", MONTANT_ALIASES),
    ("region", REGION_ALIASES),
    ("secteur", SECTEUR_ALIASES),
    ("date_depot", DATE_DEPOT_ALIASES),
    ("observations", OBSERVATIONS_ALIASES),
    ("phone", PHONE_ALIASES),
]

# Fields tracked for change detection (DossierHistory)
TRACKED_FIELDS: list[str] = [
    "statut", "type_projet", "raison_sociale", "montant_investissement",
    "region", "secteur", "date_depot", "observations",
]

PHONE_PATTERN = re.compile(r"^\+[1-9]\d{6,14}$")

# Date formats commonly used in Moroccan CRI exports
_DATE_FORMATS: list[str] = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"]

# ── Dataclasses ──────────────────────────────────────────────────────


@dataclass
class DossierImportRow:
    """Parsed row from an Excel/CSV file."""

    row_number: int
    numero: str | None = None
    statut: str | None = None
    type_projet: str | None = None
    raison_sociale: str | None = None
    montant_investissement: str | None = None
    region: str | None = None
    secteur: str | None = None
    date_depot: str | None = None
    observations: str | None = None
    phone: str | None = None
    raw_data: dict = field(default_factory=dict)


@dataclass
class FieldChange:
    """A detected change on a single dossier field."""

    field_name: str
    old_value: str | None
    new_value: str | None


@dataclass
class ImportValidationResult:
    """Result of file-level validation."""

    is_valid: bool
    file_hash: str | None = None
    file_size: int = 0
    error: str | None = None
    is_duplicate: bool = False


@dataclass
class ImportReport:
    """Detailed import report."""

    sync_log_id: uuid.UUID
    rows_total: int = 0
    rows_imported: int = 0
    rows_updated: int = 0
    rows_errored: int = 0
    errors: list[dict] = field(default_factory=list)
    duration_seconds: float = 0.0


# ── Helpers ──────────────────────────────────────────────────────────


def _normalize_header(header: str) -> str:
    """Normalize a column header for alias matching."""
    return header.strip().lower().replace(" ", "_")


def _find_column(headers: list[str], aliases: set[str]) -> int | None:
    """Return the index of the first header matching *aliases*, or None."""
    for i, h in enumerate(headers):
        if _normalize_header(h) in aliases:
            return i
    return None


def _strip_html(value: str) -> str:
    """Remove HTML tags from *value*."""
    return re.sub(r"<[^>]+>", "", value).strip()


def _cell_to_str(value: Any) -> str | None:
    """Convert an openpyxl/csv cell value to a trimmed string or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    s = str(value).strip()
    return s if s else None


def _parse_date(value: str | None) -> date | None:
    """Parse a date string using common Moroccan formats."""
    if not value:
        return None
    v = value.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def _parse_montant(value: str | None) -> Decimal | None:
    """Parse a monetary amount, stripping currency symbols."""
    if not value:
        return None
    cleaned = value.strip()
    # Strip common currency labels
    for label in ("MAD", "DH", "Dhs", "dh", "mad"):
        cleaned = cleaned.replace(label, "")
    # Normalise separators: spaces, commas
    cleaned = cleaned.replace("\u00a0", "").replace(" ", "")
    cleaned = cleaned.replace(",", ".")
    cleaned = cleaned.strip()
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


# ── Service ──────────────────────────────────────────────────────────


class DossierImportService:
    """Secure import of dossier data from Excel/CSV files."""

    def __init__(self) -> None:
        self._logger = logger.bind(service="dossier_import_service")

    # ── File validation ───────────────────────────────────────────

    def _check_extension(self, file_path: str) -> bool:
        """Return True if *file_path* has an allowed extension."""
        return PurePosixPath(file_path).suffix.lower() in ALLOWED_EXTENSIONS

    async def validate_file(
        self,
        file_path: str,
        tenant: TenantContext,
    ) -> ImportValidationResult:
        """Validate a file before import: extension, size, and duplicate hash.

        Args:
            file_path: Path to the file on disk.
            tenant: Current tenant context.

        Returns:
            ImportValidationResult with validation outcome.
        """
        # Extension check
        if not self._check_extension(file_path):
            ext = PurePosixPath(file_path).suffix.lower()
            return ImportValidationResult(
                is_valid=False,
                error=f"Extension non autorisée : {ext}. "
                       f"Formats acceptés : {', '.join(sorted(ALLOWED_EXTENSIONS))}",
            )

        # Size check
        try:
            file_size = os.path.getsize(file_path)
        except OSError as exc:
            return ImportValidationResult(
                is_valid=False,
                error=f"Fichier inaccessible : {exc}",
            )

        max_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
        if file_size > max_bytes:
            return ImportValidationResult(
                is_valid=False,
                file_size=file_size,
                error=f"Fichier trop volumineux : {file_size / 1024 / 1024:.1f} Mo "
                       f"(max {MAX_FILE_SIZE_MB} Mo)",
            )

        # SHA-256 hash (read in 8 KB chunks to limit memory)
        sha = hashlib.sha256()
        with open(file_path, "rb") as fh:
            while True:
                chunk = fh.read(8192)
                if not chunk:
                    break
                sha.update(chunk)
        file_hash = sha.hexdigest()

        # Duplicate check against completed sync_logs
        async with tenant.db_session() as session:
            result = await session.execute(
                select(SyncLog.id).where(
                    SyncLog.file_hash == file_hash,
                    SyncLog.status == SyncStatus.completed,
                ),
            )
            if result.scalar_one_or_none() is not None:
                return ImportValidationResult(
                    is_valid=False,
                    file_hash=file_hash,
                    file_size=file_size,
                    is_duplicate=True,
                    error="Ce fichier a déjà été importé (hash SHA-256 identique).",
                )

        return ImportValidationResult(
            is_valid=True,
            file_hash=file_hash,
            file_size=file_size,
        )

    # ── Excel parsing ─────────────────────────────────────────────

    def parse_excel(
        self,
        file_path: str,
        column_mapping: dict[str, str] | None = None,
    ) -> list[DossierImportRow]:
        """Parse an Excel file into import rows.

        Args:
            file_path: Path to .xlsx/.xls file.
            column_mapping: Optional explicit mapping ``{source_header: field_name}``.
                If *None*, columns are auto-detected via alias sets.

        Returns:
            List of parsed DossierImportRow.
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return []

        rows_iter = ws.iter_rows()

        # Read headers from first row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            return []

        raw_headers = [_cell_to_str(cell.value) or "" for cell in header_row]
        col_map = self._build_column_map(raw_headers, column_mapping)

        results: list[DossierImportRow] = []
        for row_idx, row in enumerate(rows_iter, start=2):
            if len(results) >= MAX_ROWS:
                self._logger.warning(
                    "max_rows_reached", max_rows=MAX_ROWS, file=file_path,
                )
                break

            values = [_cell_to_str(cell.value) for cell in row]

            # Skip entirely empty rows
            if not any(values):
                continue

            raw_data = {
                raw_headers[i]: values[i]
                for i in range(min(len(raw_headers), len(values)))
                if values[i] is not None
            }

            import_row = DossierImportRow(row_number=row_idx, raw_data=raw_data)
            for field_name, col_idx in col_map.items():
                if col_idx < len(values):
                    setattr(import_row, field_name, values[col_idx])
            results.append(import_row)

        wb.close()
        return results

    # ── CSV parsing ───────────────────────────────────────────────

    def parse_csv(
        self,
        file_path: str,
        column_mapping: dict[str, str] | None = None,
    ) -> list[DossierImportRow]:
        """Parse a CSV file into import rows.

        Detects encoding (UTF-8-sig → UTF-8 → Windows-1252 → Latin-1)
        and separator (Sniffer → ``;`` → ``,``).

        Args:
            file_path: Path to .csv file.
            column_mapping: Optional explicit column mapping.

        Returns:
            List of parsed DossierImportRow.
        """
        raw_bytes = open(file_path, "rb").read()
        text = self._decode_csv(raw_bytes)
        delimiter = self._detect_delimiter(text)

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows_iter = iter(reader)

        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            return []

        col_map = self._build_column_map(raw_headers, column_mapping)

        results: list[DossierImportRow] = []
        for row_idx, raw_row in enumerate(rows_iter, start=2):
            if len(results) >= MAX_ROWS:
                self._logger.warning(
                    "max_rows_reached", max_rows=MAX_ROWS, file=file_path,
                )
                break

            if not any(cell.strip() for cell in raw_row):
                continue

            values = [v.strip() if v else None for v in raw_row]
            raw_data = {
                raw_headers[i]: values[i]
                for i in range(min(len(raw_headers), len(values)))
                if values[i] is not None
            }

            import_row = DossierImportRow(row_number=row_idx, raw_data=raw_data)
            for field_name, col_idx in col_map.items():
                if col_idx < len(values):
                    setattr(import_row, field_name, values[col_idx])
            results.append(import_row)

        return results

    # ── Sanitisation ──────────────────────────────────────────────

    def _sanitize_string(self, value: str | None) -> str | None:
        """Strip whitespace, HTML tags, and neutralise SQL injection patterns."""
        if value is None:
            return None
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = _strip_html(cleaned)
        for pattern in SQL_INJECTION_PATTERNS:
            cleaned = pattern.sub("", cleaned)
        return cleaned.strip() or None

    def _normalize_moroccan_phone(self, raw: str | None) -> str | None:
        """Normalise a Moroccan phone number to E.164 format.

        Handles local (06/07), 00212, and +212 formats.
        Returns None for unrecognisable numbers.
        """
        if not raw:
            return None
        # Strip whitespace, dashes, dots, parentheses
        phone = raw.strip().replace(" ", "").replace("-", "").replace(".", "")
        phone = phone.replace("(", "").replace(")", "")

        # TODO ─────────────────────────────────────────────────────
        # This is your part to implement! The method should:
        #
        # 1. Handle local format: 06XXXXXXXX or 07XXXXXXXX (10 digits)
        #    → prepend +212 and drop the leading 0
        # 2. Handle 00212XXXXXXXXX → replace 00 with +
        # 3. Handle 212XXXXXXXXX (12 digits, no prefix) → prepend +
        # 4. Keep numbers already starting with +212 as-is
        # 5. Keep other international numbers (+33..., +1...) as-is
        # 6. Validate final result matches E.164: ^\+[1-9]\d{6,14}$
        # 7. Return None if the number can't be normalised
        #
        # Consider:
        # - Moroccan mobile starts with 06 or 07 (9 digits after +212)
        # - Moroccan landline starts with 05 (also valid)
        # - What about numbers with fewer digits? (invalid → None)
        #
        # Implement below, replacing the placeholder:
        # ──────────────────────────────────────────────────────────

        # Local format: 0XXXXXXXXX (10 digits)
        if phone.startswith("0") and len(phone) == 10:
            phone = "+212" + phone[1:]
        # 00212 prefix
        elif phone.startswith("00212") and len(phone) == 14:
            phone = "+" + phone[2:]
        # 212 without + prefix (12 digits)
        elif phone.startswith("212") and len(phone) == 12:
            phone = "+" + phone
        # Already +212 or other international
        elif phone.startswith("+"):
            pass
        else:
            return None

        # Final E.164 validation
        if not PHONE_PATTERN.match(phone):
            return None
        return phone

    def sanitize_row(self, row: DossierImportRow) -> DossierImportRow:
        """Sanitise all fields in an import row.

        Applies HTML stripping, SQL injection neutralisation,
        phone normalisation, and whitespace trimming.

        Args:
            row: Raw parsed row.

        Returns:
            The same row with sanitised field values.
        """
        row.numero = self._sanitize_string(row.numero)
        row.statut = self._sanitize_string(row.statut)
        row.type_projet = self._sanitize_string(row.type_projet)
        row.raison_sociale = self._sanitize_string(row.raison_sociale)
        row.montant_investissement = self._sanitize_string(
            row.montant_investissement,
        )
        row.region = self._sanitize_string(row.region)
        row.secteur = self._sanitize_string(row.secteur)
        row.observations = self._sanitize_string(row.observations)
        row.date_depot = self._sanitize_string(row.date_depot)
        row.phone = self._normalize_moroccan_phone(row.phone)
        return row

    # ── Change detection ──────────────────────────────────────────

    def detect_changes(
        self,
        existing: Any,  # Dossier ORM instance
        new_data: DossierImportRow,
    ) -> list[FieldChange]:
        """Compare an existing dossier against new import data.

        Only fields with a non-None new value that differ from the existing
        value produce a FieldChange entry.

        Args:
            existing: Existing Dossier ORM object.
            new_data: New data from the import row.

        Returns:
            List of detected field changes.
        """
        changes: list[FieldChange] = []

        for field_name in TRACKED_FIELDS:
            new_raw = getattr(new_data, field_name, None)
            if new_raw is None:
                continue

            existing_val = getattr(existing, field_name, None)

            # Convert both sides to comparable strings
            if field_name == "statut":
                new_mapped = STATUT_MAPPING.get(
                    new_raw.strip().lower(), DossierStatut.en_attente,
                )
                new_str = new_mapped.value
                old_str = existing_val.value if existing_val else None
            elif field_name == "montant_investissement":
                new_parsed = _parse_montant(new_raw)
                new_str = str(new_parsed) if new_parsed is not None else None
                old_str = str(existing_val) if existing_val is not None else None
            elif field_name == "date_depot":
                new_parsed_date = _parse_date(new_raw)
                new_str = str(new_parsed_date) if new_parsed_date else None
                old_str = str(existing_val) if existing_val else None
            else:
                new_str = new_raw.strip() if new_raw else None
                old_str = str(existing_val) if existing_val is not None else None

            if new_str is None:
                continue
            if old_str == new_str:
                continue

            changes.append(FieldChange(
                field_name=field_name,
                old_value=old_str,
                new_value=new_str,
            ))

        return changes

    # ── Main import logic ─────────────────────────────────────────

    async def import_dossiers(
        self,
        rows: list[DossierImportRow],
        sync_log_id: uuid.UUID,
        tenant: TenantContext,
    ) -> ImportReport:
        """Import sanitised rows into the tenant's dossier table.

        Performs upsert by ``numero``, detects changes, creates history
        entries, and links contacts by phone number.

        Args:
            rows: Pre-sanitised import rows.
            sync_log_id: ID of the SyncLog tracking this import.
            tenant: Current tenant context.

        Returns:
            ImportReport with detailed statistics.
        """
        start = time.monotonic()
        report = ImportReport(sync_log_id=sync_log_id, rows_total=len(rows))

        async with tenant.db_session() as session:
            # Mark SyncLog as running
            sync_log_result = await session.execute(
                select(SyncLog).where(SyncLog.id == sync_log_id),
            )
            sync_log = sync_log_result.scalar_one_or_none()
            if sync_log:
                sync_log.status = SyncStatus.running
                sync_log.started_at = datetime.utcnow()
                sync_log.rows_total = len(rows)
                await session.flush()

            # Process in batches
            for batch_start in range(0, len(rows), BATCH_SIZE):
                batch = rows[batch_start : batch_start + BATCH_SIZE]
                await self._process_batch(batch, sync_log_id, session, report)

            # Finalise SyncLog
            if sync_log:
                sync_log.status = SyncStatus.completed
                sync_log.completed_at = datetime.utcnow()
                sync_log.rows_imported = report.rows_imported
                sync_log.rows_updated = report.rows_updated
                sync_log.rows_errored = report.rows_errored
                if report.errors:
                    sync_log.error_details = {"errors": report.errors}

        report.duration_seconds = round(time.monotonic() - start, 3)

        self._logger.info(
            "import_completed",
            tenant=tenant.slug,
            sync_log_id=str(sync_log_id),
            total=report.rows_total,
            imported=report.rows_imported,
            updated=report.rows_updated,
            errored=report.rows_errored,
            duration_s=report.duration_seconds,
        )
        return report

    # ── Internal helpers ──────────────────────────────────────────

    async def _process_batch(
        self,
        batch: list[DossierImportRow],
        sync_log_id: uuid.UUID,
        session: Any,
        report: ImportReport,
    ) -> None:
        """Process a batch of import rows within an open session."""
        # Batch-lookup existing dossiers by numero
        numeros = [r.numero for r in batch if r.numero]
        existing_map: dict[str, Any] = {}
        if numeros:
            result = await session.execute(
                select(Dossier).where(Dossier.numero.in_(numeros)),
            )
            for dossier in result.scalars().all():
                existing_map[dossier.numero] = dossier

        # Batch-lookup contacts by phone
        phones = [r.phone for r in batch if r.phone]
        phone_to_contact: dict[str, uuid.UUID] = {}
        if phones:
            result = await session.execute(
                select(Contact.id, Contact.phone).where(
                    Contact.phone.in_(phones),
                ),
            )
            for contact_id, phone in result.all():
                phone_to_contact[phone] = contact_id

        # Process each row
        for row in batch:
            try:
                if not row.numero:
                    report.rows_errored += 1
                    report.errors.append({
                        "row": row.row_number,
                        "field": "numero",
                        "error": "Numéro de dossier manquant",
                    })
                    continue

                existing = existing_map.get(row.numero)
                if existing:
                    self._update_dossier(
                        existing, row, sync_log_id, phone_to_contact, session,
                        report,
                    )
                else:
                    self._create_dossier(
                        row, sync_log_id, phone_to_contact, session, report,
                    )
            except Exception as exc:
                self._logger.error(
                    "row_import_error",
                    row=row.row_number,
                    numero=row.numero,
                    error=str(exc),
                    exc_info=True,
                )
                report.rows_errored += 1
                report.errors.append({
                    "row": row.row_number,
                    "field": "general",
                    "error": str(exc),
                })

        await session.flush()

    def _create_dossier(
        self,
        row: DossierImportRow,
        sync_log_id: uuid.UUID,
        phone_to_contact: dict[str, uuid.UUID],
        session: Any,
        report: ImportReport,
    ) -> None:
        """Create a new Dossier from an import row."""
        statut = DossierStatut.en_attente
        if row.statut:
            statut = STATUT_MAPPING.get(
                row.statut.strip().lower(), DossierStatut.en_attente,
            )

        dossier = Dossier(
            numero=row.numero,
            statut=statut,
            type_projet=row.type_projet,
            raison_sociale=row.raison_sociale,
            montant_investissement=_parse_montant(row.montant_investissement),
            region=row.region,
            secteur=row.secteur,
            date_depot=_parse_date(row.date_depot),
            date_derniere_maj=date.today(),
            observations=row.observations,
            raw_data=row.raw_data or None,
        )

        # Link contact by phone
        if row.phone and row.phone in phone_to_contact:
            dossier.contact_id = phone_to_contact[row.phone]

        session.add(dossier)
        report.rows_imported += 1

    def _update_dossier(
        self,
        existing: Any,
        row: DossierImportRow,
        sync_log_id: uuid.UUID,
        phone_to_contact: dict[str, uuid.UUID],
        session: Any,
        report: ImportReport,
    ) -> None:
        """Update an existing Dossier and create history entries."""
        changes = self.detect_changes(existing, row)

        if not changes:
            return

        for change in changes:
            history = DossierHistory(
                dossier_id=existing.id,
                field_changed=change.field_name,
                old_value=change.old_value,
                new_value=change.new_value,
                sync_log_id=sync_log_id,
            )
            session.add(history)

        # Apply changes to dossier
        for change in changes:
            if change.field_name == "statut":
                existing.statut = STATUT_MAPPING.get(
                    (row.statut or "").strip().lower(),
                    existing.statut,
                )
            elif change.field_name == "montant_investissement":
                existing.montant_investissement = _parse_montant(
                    row.montant_investissement,
                )
            elif change.field_name == "date_depot":
                existing.date_depot = _parse_date(row.date_depot)
            else:
                setattr(existing, change.field_name, change.new_value)

        existing.date_derniere_maj = date.today()

        # Update raw_data
        if row.raw_data:
            existing.raw_data = row.raw_data

        # Link contact if not already linked
        if row.phone and row.phone in phone_to_contact and not existing.contact_id:
            existing.contact_id = phone_to_contact[row.phone]

        report.rows_updated += 1

    def _build_column_map(
        self,
        raw_headers: list[str],
        column_mapping: dict[str, str] | None,
    ) -> dict[str, int]:
        """Build a {field_name: column_index} mapping.

        If *column_mapping* is provided (``{source_header: field_name}``),
        use it. Otherwise, auto-detect via alias sets.
        """
        if column_mapping:
            result: dict[str, int] = {}
            norm_headers = [_normalize_header(h) for h in raw_headers]
            for source_header, field_name in column_mapping.items():
                norm = _normalize_header(source_header)
                if norm in norm_headers:
                    result[field_name] = norm_headers.index(norm)
            return result

        # Auto-detect via alias sets
        col_map: dict[str, int] = {}
        for field_name, aliases in _FIELD_ALIAS_MAP:
            idx = _find_column(raw_headers, aliases)
            if idx is not None:
                col_map[field_name] = idx
        return col_map

    @staticmethod
    def _decode_csv(raw_bytes: bytes) -> str:
        """Decode CSV bytes, trying common encodings."""
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                return raw_bytes.decode(encoding)
            except (UnicodeDecodeError, ValueError):
                continue
        # latin-1 never raises, so this is a safety fallback
        return raw_bytes.decode("latin-1")

    @staticmethod
    def _detect_delimiter(text: str) -> str:
        """Detect CSV delimiter using Sniffer, with fallbacks."""
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            return dialect.delimiter
        except csv.Error:
            # Fallback: count occurrences in first line
            first_line = sample.split("\n", 1)[0]
            if first_line.count(";") > first_line.count(","):
                return ";"
            return ","


# ── Singleton ────────────────────────────────────────────────────────

_dossier_import_service: DossierImportService | None = None


def get_dossier_import_service() -> DossierImportService:
    """Return the singleton DossierImportService instance."""
    global _dossier_import_service
    if _dossier_import_service is None:
        _dossier_import_service = DossierImportService()
    return _dossier_import_service
