"""ContactImportExportService — import contacts from Excel/CSV, export to Excel/CSV.

Import validates phone format, deduplicates by phone number, and bulk-inserts.
Export generates files using openpyxl (xlsx) or csv module.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import PurePosixPath

import structlog
from openpyxl import Workbook, load_workbook
from sqlalchemy import Select, or_, select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core.exceptions import ValidationError
from app.core.tenant import TenantContext
from app.models.contact import Contact
from app.models.enums import ContactSource, Language, OptInStatus

logger = structlog.get_logger()

PHONE_PATTERN = re.compile(r"^\+[1-9]\d{6,14}$")
CIN_PATTERN = re.compile(r"^[A-Z]{1,2}\d{5,6}$")
MAX_IMPORT_ROWS = 50_000
BATCH_SIZE = 500

# Column aliases for flexible CSV/Excel headers
PHONE_ALIASES = {"phone", "telephone", "téléphone", "tel", "mobile", "numéro"}
NAME_ALIASES = {"name", "nom", "full_name", "nom_complet"}
LANGUAGE_ALIASES = {"language", "langue", "lang"}
CIN_ALIASES = {"cin", "id_card", "carte_identité"}
TAGS_ALIASES = {"tags", "étiquettes", "labels"}

VALID_LANGUAGES = {lang.value for lang in Language}


@dataclass
class ImportResult:
    """Result of a contact import operation."""

    created: int = 0
    skipped: int = 0
    errors: list[dict] = field(default_factory=list)


def _normalize_header(header: str) -> str:
    """Normalize a column header for matching."""
    return header.strip().lower().replace(" ", "_")


def _find_column(headers: list[str], aliases: set[str]) -> int | None:
    """Find the index of a column matching one of the aliases."""
    for i, h in enumerate(headers):
        if _normalize_header(h) in aliases:
            return i
    return None


def _strip_html(value: str) -> str:
    """Strip HTML tags from a string value."""
    return re.sub(r"<[^>]+>", "", value).strip()


class ContactImportExportService:
    """Import and export contacts for the back-office."""

    def __init__(self) -> None:
        self._logger = logger.bind(service="contact_import_export")

    async def import_contacts(
        self,
        tenant: TenantContext,
        file_bytes: bytes,
        filename: str,
    ) -> ImportResult:
        """Import contacts from an Excel or CSV file.

        Steps:
        1. Detect format from filename extension
        2. Parse rows
        3. Validate phone, normalize fields
        4. Dedup against existing contacts in tenant
        5. Bulk insert new contacts

        Args:
            tenant: Tenant context for DB session.
            file_bytes: Raw file content.
            filename: Original filename (for extension detection).

        Returns:
            ImportResult with created/skipped/errors counts.

        Raises:
            ValidationError: If file format unsupported or too many rows.
        """
        ext = PurePosixPath(filename).suffix.lower()
        if ext == ".csv":
            rows = self._parse_csv(file_bytes)
        elif ext in (".xlsx", ".xls"):
            rows = self._parse_excel(file_bytes)
        else:
            raise ValidationError(
                f"Unsupported file format: {ext}. Use .csv or .xlsx",
                details={"extension": ext},
            )

        if len(rows) > MAX_IMPORT_ROWS:
            raise ValidationError(
                f"Too many rows: {len(rows)} (max {MAX_IMPORT_ROWS})",
                details={"row_count": len(rows), "max": MAX_IMPORT_ROWS},
            )

        self._logger.info(
            "import_started",
            tenant=tenant.slug,
            filename=filename,
            row_count=len(rows),
        )

        result = ImportResult()

        # Validate rows
        valid_rows: list[dict] = []
        for idx, row in enumerate(rows, start=2):  # Row 2+ (1 is header)
            phone = row.get("phone", "").strip()
            if not phone:
                result.errors.append({"row": idx, "phone": None, "error": "Phone manquant"})
                continue

            # Normalize phone: add + if missing
            if not phone.startswith("+"):
                phone = f"+{phone}"

            if not PHONE_PATTERN.match(phone):
                result.errors.append(
                    {"row": idx, "phone": phone, "error": "Format téléphone invalide (E.164)"}
                )
                continue

            name = _strip_html(row.get("name", "") or "").strip() or None
            lang_str = (row.get("language", "") or "").strip().lower()
            language = lang_str if lang_str in VALID_LANGUAGES else Language.fr.value
            cin = (row.get("cin", "") or "").strip().upper() or None

            if cin and not CIN_PATTERN.match(cin):
                result.errors.append(
                    {"row": idx, "phone": phone, "error": f"Format CIN invalide: {cin}"}
                )
                continue

            tags_raw = (row.get("tags", "") or "").strip()
            tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

            valid_rows.append(
                {
                    "phone": phone,
                    "name": name,
                    "language": language,
                    "cin": cin,
                    "tags": tags,
                    "source": ContactSource.import_csv.value,
                    "opt_in_status": OptInStatus.pending.value,
                }
            )

        if not valid_rows:
            return result

        # Batch dedup and insert
        async with tenant.db_session() as session:
            for i in range(0, len(valid_rows), BATCH_SIZE):
                batch = valid_rows[i : i + BATCH_SIZE]
                batch_phones = [r["phone"] for r in batch]

                # Find existing phones
                existing_result = await session.execute(
                    select(Contact.phone).where(Contact.phone.in_(batch_phones)),
                )
                existing_phones = {row[0] for row in existing_result.all()}

                # Split into new and existing
                new_rows = [r for r in batch if r["phone"] not in existing_phones]
                result.skipped += len(batch) - len(new_rows)

                if new_rows:
                    await session.execute(
                        pg_insert(Contact).values(new_rows),
                    )
                    result.created += len(new_rows)

        self._logger.info(
            "import_completed",
            tenant=tenant.slug,
            created=result.created,
            skipped=result.skipped,
            errors=len(result.errors),
        )
        return result

    def _parse_csv(self, file_bytes: bytes) -> list[dict]:
        """Parse CSV file bytes into a list of dicts with normalized keys."""
        # Try UTF-8, fall back to latin-1
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")

        reader = csv.reader(io.StringIO(text))
        rows_iter = iter(reader)

        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            return []

        # Map columns
        phone_idx = _find_column(raw_headers, PHONE_ALIASES)
        name_idx = _find_column(raw_headers, NAME_ALIASES)
        lang_idx = _find_column(raw_headers, LANGUAGE_ALIASES)
        cin_idx = _find_column(raw_headers, CIN_ALIASES)
        tags_idx = _find_column(raw_headers, TAGS_ALIASES)

        if phone_idx is None:
            raise ValidationError(
                "Colonne 'phone' introuvable dans le CSV",
                details={"headers": raw_headers},
            )

        rows = []
        for raw_row in rows_iter:
            if not any(cell.strip() for cell in raw_row):
                continue
            rows.append(
                {
                    "phone": raw_row[phone_idx] if phone_idx < len(raw_row) else "",
                    "name": raw_row[name_idx]
                    if name_idx is not None and name_idx < len(raw_row)
                    else "",
                    "language": raw_row[lang_idx]
                    if lang_idx is not None and lang_idx < len(raw_row)
                    else "",
                    "cin": raw_row[cin_idx]
                    if cin_idx is not None and cin_idx < len(raw_row)
                    else "",
                    "tags": raw_row[tags_idx]
                    if tags_idx is not None and tags_idx < len(raw_row)
                    else "",
                }
            )
        return rows

    def _parse_excel(self, file_bytes: bytes) -> list[dict]:
        """Parse Excel file bytes into a list of dicts with normalized keys."""
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)

        try:
            raw_headers = [str(h or "").strip() for h in next(rows_iter)]
        except StopIteration:
            wb.close()
            return []

        phone_idx = _find_column(raw_headers, PHONE_ALIASES)
        name_idx = _find_column(raw_headers, NAME_ALIASES)
        lang_idx = _find_column(raw_headers, LANGUAGE_ALIASES)
        cin_idx = _find_column(raw_headers, CIN_ALIASES)
        tags_idx = _find_column(raw_headers, TAGS_ALIASES)

        if phone_idx is None:
            wb.close()
            raise ValidationError(
                "Colonne 'phone' introuvable dans le fichier Excel",
                details={"headers": raw_headers},
            )

        rows = []
        for raw_row in rows_iter:
            cells = [str(c or "").strip() if c is not None else "" for c in raw_row]
            if not any(cells):
                continue
            rows.append(
                {
                    "phone": cells[phone_idx] if phone_idx < len(cells) else "",
                    "name": cells[name_idx]
                    if name_idx is not None and name_idx < len(cells)
                    else "",
                    "language": cells[lang_idx]
                    if lang_idx is not None and lang_idx < len(cells)
                    else "",
                    "cin": cells[cin_idx] if cin_idx is not None and cin_idx < len(cells) else "",
                    "tags": cells[tags_idx]
                    if tags_idx is not None and tags_idx < len(cells)
                    else "",
                }
            )
        wb.close()
        return rows

    # ── Filtered export (Wave 17) ──

    @staticmethod
    def _build_filtered_query(
        *,
        search: str | None = None,
        opt_in_status: OptInStatus | None = None,
        language: Language | None = None,
        tags: list[str] | None = None,
        source: ContactSource | None = None,
        created_after: datetime | None = None,
        created_before: datetime | None = None,
    ) -> Select[tuple[Contact]]:
        """Build a filtered ``select(Contact)`` for export.

        When all params are None the query returns all contacts (backward
        compatible with pre-Wave-17 behaviour).
        """
        base = select(Contact)
        if search:
            pattern = f"%{search}%"
            base = base.where(
                or_(
                    Contact.name.ilike(pattern),
                    Contact.phone.ilike(pattern),
                    Contact.cin.ilike(pattern),
                ),
            )
        if opt_in_status is not None:
            base = base.where(Contact.opt_in_status == opt_in_status)
        if language is not None:
            base = base.where(Contact.language == language)
        if tags:
            for tag in tags:
                base = base.where(Contact.tags.contains([tag]))
        if source is not None:
            base = base.where(Contact.source == source)
        if created_after is not None:
            base = base.where(Contact.created_at >= created_after)
        if created_before is not None:
            base = base.where(Contact.created_at <= created_before)
        return base.order_by(Contact.created_at.desc())

    async def export_to_xlsx(
        self,
        tenant: TenantContext,
        *,
        search: str | None = None,
        opt_in_status: OptInStatus | None = None,
        language: Language | None = None,
        tags: list[str] | None = None,
        source: ContactSource | None = None,
        created_after: datetime | None = None,
        created_before: datetime | None = None,
    ) -> bytes:
        """Export contacts to Excel bytes, optionally filtered."""
        query = self._build_filtered_query(
            search=search,
            opt_in_status=opt_in_status,
            language=language,
            tags=tags,
            source=source,
            created_after=created_after,
            created_before=created_before,
        )
        async with tenant.db_session() as session:
            result = await session.execute(query)
            contacts = result.scalars().all()

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Contacts")
        ws.append(["Téléphone", "Nom", "Langue", "CIN", "Opt-in", "Tags", "Source", "Créé le"])

        for c in contacts:
            ws.append(
                [
                    c.phone,
                    c.name or "",
                    c.language.value if c.language else "",
                    c.cin or "",
                    c.opt_in_status.value if c.opt_in_status else "",
                    ", ".join(c.tags) if c.tags else "",
                    c.source.value if c.source else "",
                    c.created_at.isoformat() if c.created_at else "",
                ]
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def export_to_csv(
        self,
        tenant: TenantContext,
        *,
        search: str | None = None,
        opt_in_status: OptInStatus | None = None,
        language: Language | None = None,
        tags: list[str] | None = None,
        source: ContactSource | None = None,
        created_after: datetime | None = None,
        created_before: datetime | None = None,
    ) -> str:
        """Export contacts to CSV string, optionally filtered."""
        query = self._build_filtered_query(
            search=search,
            opt_in_status=opt_in_status,
            language=language,
            tags=tags,
            source=source,
            created_after=created_after,
            created_before=created_before,
        )
        async with tenant.db_session() as session:
            result = await session.execute(query)
            contacts = result.scalars().all()

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            ["phone", "name", "language", "cin", "opt_in_status", "tags", "source", "created_at"]
        )

        for c in contacts:
            writer.writerow(
                [
                    c.phone,
                    c.name or "",
                    c.language.value if c.language else "",
                    c.cin or "",
                    c.opt_in_status.value if c.opt_in_status else "",
                    ", ".join(c.tags) if c.tags else "",
                    c.source.value if c.source else "",
                    c.created_at.isoformat() if c.created_at else "",
                ]
            )

        return buffer.getvalue()


# ---------------------------------------------------------------------------
# Singleton
# ---------------------------------------------------------------------------
_import_export_service: ContactImportExportService | None = None


def get_import_export_service() -> ContactImportExportService:
    """Get or create the ContactImportExportService singleton."""
    global _import_export_service  # noqa: PLW0603
    if _import_export_service is None:
        _import_export_service = ContactImportExportService()
    return _import_export_service
